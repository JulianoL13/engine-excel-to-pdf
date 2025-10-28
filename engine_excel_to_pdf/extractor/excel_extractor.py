from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING, List, Optional

if TYPE_CHECKING:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet

from ..models import Certificado, CertificadoBundle, MetodoAplicacao, ProdutoQuimico
from ..utils import normalize_whitespace, parse_pt_br_date


@dataclass(slots=True)
class ExcelExtractorConfig:
    certificado_map: dict[str, str]
    classe_quimica_column: int
    classe_quimica_start_row: int
    produto_nome_column: int
    produto_concentracao_column: int
    produto_start_row: int
    metodo_descricao_column: int
    metodo_quantidade_column: int
    metodo_start_row: int
    metodo_end_row: Optional[int] = None


DEFAULT_CONFIG = ExcelExtractorConfig(
    certificado_map={
        "numero_certificado": "C10",
        "numero_licenca": "I10",
        "razao_social": "C15",
        "nome_fantasia": "C17",
        "endereco_completo": "D19",
        "cnpj": "E20",
        "data_execucao": "E21",
        "pragas_tratadas": "F22",
        "data_validade": "B48",
    },
    classe_quimica_column=6,
    classe_quimica_start_row=25,
    produto_nome_column=4,
    produto_concentracao_column=9,
    produto_start_row=29,
    metodo_descricao_column=4,
    metodo_quantidade_column=8,
    metodo_start_row=34,
    metodo_end_row=80,
)


class ExcelExtractor:
    def __init__(self, config: ExcelExtractorConfig = DEFAULT_CONFIG):
        self.config = config

    def extract(self, file_path: Path) -> CertificadoBundle:
        from openpyxl import load_workbook
        
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active

        certificado_data = self._extract_certificado(worksheet, file_path.name)
        produtos = self._extract_produtos(worksheet)
        metodos = self._extract_metodos(worksheet)

        return CertificadoBundle(certificado=certificado_data, produtos=produtos, metodos=metodos)

    def _extract_certificado(self, worksheet: "Worksheet", arquivo_origem: str) -> Certificado:
        values: dict[str, str] = {}
        for field, cell_ref in self.config.certificado_map.items():
            cell_value = worksheet[cell_ref].value if worksheet[cell_ref].value is not None else ""
            values[field] = normalize_whitespace(str(cell_value))

        data_execucao = parse_pt_br_date(values["data_execucao"])
        data_validade = parse_pt_br_date(values["data_validade"])

        endereco_completo = values["endereco_completo"]
        bairro = None
        cidade = None
        
        if endereco_completo:
            partes = [p.strip() for p in endereco_completo.split(",")]
            if len(partes) >= 3:
                bairro = partes[-2]
                cidade = partes[-1]

        return Certificado(
            numero_certificado=values["numero_certificado"],
            numero_licenca=values["numero_licenca"],
            razao_social=values["razao_social"],
            nome_fantasia=values["nome_fantasia"],
            cnpj=values["cnpj"],
            endereco_completo=endereco_completo,
            data_execucao=data_execucao,
            data_validade=data_validade,
            pragas_tratadas=values["pragas_tratadas"],
            arquivo_origem=arquivo_origem,
            data_cadastro=datetime.now(timezone.utc),
            bairro=bairro,
            cidade=cidade,
        )

    def _extract_produtos(self, worksheet: "Worksheet") -> List[ProdutoQuimico]:
        produtos: List[ProdutoQuimico] = []
        
        classes = self._extract_column_values(
            worksheet,
            column=self.config.classe_quimica_column,
            start_row=self.config.classe_quimica_start_row
        )

        row = self.config.produto_start_row
        index = 0
        while True:
            nome = worksheet.cell(row=row, column=self.config.produto_nome_column).value
            if not nome or not str(nome).strip():
                break

            concentracao_value = worksheet.cell(row=row, column=self.config.produto_concentracao_column).value
            concentracao = self._convert_concentracao(concentracao_value)
            classe = classes[index] if index < len(classes) else ""

            produtos.append(
                ProdutoQuimico(
                    nome_produto=normalize_whitespace(str(nome)),
                    classe_quimica=normalize_whitespace(classe),
                    concentracao=concentracao,
                )
            )
            index += 1
            row += 1

        return produtos

    def _extract_metodos(self, worksheet: "Worksheet") -> List[MetodoAplicacao]:
        metodos: List[MetodoAplicacao] = []
        start = self.config.metodo_start_row
        end = self.config.metodo_end_row or worksheet.max_row

        for row in range(start, end + 1):
            descricao = worksheet.cell(row=row, column=self.config.metodo_descricao_column).value
            if not descricao or not str(descricao).strip():
                continue
            quantidade = worksheet.cell(row=row, column=self.config.metodo_quantidade_column).value
            quantidade_str = normalize_whitespace(str(quantidade)) if quantidade not in (None, "") else ""
            metodos.append(
                MetodoAplicacao(
                    metodo=normalize_whitespace(str(descricao)),
                    quantidade=quantidade_str,
                )
            )

        return metodos

    def _extract_column_values(
        self, worksheet: "Worksheet", column: int, start_row: int, end_row: Optional[int] = None
    ) -> List[str]:
        values: List[str] = []
        row = start_row
        max_row = end_row or worksheet.max_row

        while row <= max_row:
            cell_value = worksheet.cell(row=row, column=column).value
            if not cell_value or not str(cell_value).strip():
                if end_row is None:
                    break
                row += 1
                continue
            values.append(normalize_whitespace(str(cell_value)))
            row += 1

        return values

    @staticmethod
    def _convert_concentracao(value) -> Optional[float]:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        value_str = str(value).strip().replace(",", ".")
        try:
            return float(value_str)
        except ValueError:
            return None
