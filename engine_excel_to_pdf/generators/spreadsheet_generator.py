from __future__ import annotations

import threading
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

from ..models import Certificado, CertificadoBundle, MetodoAplicacao, ProdutoQuimico
from ..config_defaults import PLANILHAS_DIR, ensure_directories
from ..utils import generate_unique_filename
from ..constants import (
    EXCEL_COLUMN_WIDTH_DEFAULT,
    EXCEL_COLUMN_WIDTH_WIDE,
    EXCEL_FREEZE_PANES_CELL,
)


class SpreadsheetGenerator:
    _lock = threading.Lock()
    
    def __init__(self, output_dir: Path = PLANILHAS_DIR, consolidated_filename: str = "certificados_consolidados.xlsx"):
        ensure_directories()
        self.output_dir = output_dir
        self.consolidated_path = self.output_dir / consolidated_filename

    def generate(self, bundle: CertificadoBundle) -> Path:
        with self._lock:
            return self._generate_unsafe(bundle)
    
    def _generate_unsafe(self, bundle: CertificadoBundle) -> Path:
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter
        
        certificado = bundle.certificado
        
        if self.consolidated_path.exists():
            workbook = load_workbook(self.consolidated_path)
            certificado_sheet = workbook["certificado"]
            produtos_sheet = workbook["produtos"]
            metodos_sheet = workbook["metodos"]
        else:
            workbook = Workbook()
            certificado_sheet = workbook.active
            certificado_sheet.title = "certificado"
            
            cert_headers = [
                "id",
                "numero_certificado",
                "numero_licenca",
                "razao_social",
                "nome_fantasia",
                "cnpj",
                "endereco_completo",
                "data_execucao",
                "data_validade",
                "pragas_tratadas",
                "arquivo_origem",
                "data_cadastro",
                "valor",
                "bairro",
                "cidade",
            ]
            certificado_sheet.append(cert_headers)
            
            for column in range(1, len(cert_headers) + 1):
                certificado_sheet.column_dimensions[get_column_letter(column)].width = EXCEL_COLUMN_WIDTH_DEFAULT
            certificado_sheet.freeze_panes = EXCEL_FREEZE_PANES_CELL
            
            produtos_sheet = workbook.create_sheet("produtos")
            produtos_sheet.append(["numero_certificado", "nome_produto", "classe_quimica", "concentracao"])
            produtos_sheet.freeze_panes = EXCEL_FREEZE_PANES_CELL
            for column in range(1, 5):
                produtos_sheet.column_dimensions[get_column_letter(column)].width = EXCEL_COLUMN_WIDTH_WIDE
            
            metodos_sheet = workbook.create_sheet("metodos")
            metodos_sheet.append(["numero_certificado", "metodo", "quantidade"])
            metodos_sheet.freeze_panes = EXCEL_FREEZE_PANES_CELL
            for column in range(1, 4):
                metodos_sheet.column_dimensions[get_column_letter(column)].width = EXCEL_COLUMN_WIDTH_WIDE
        
        certificado_sheet.append(
            [
                certificado.id or "",
                certificado.numero_certificado,
                certificado.numero_licenca,
                certificado.razao_social,
                certificado.nome_fantasia,
                certificado.cnpj,
                certificado.endereco_completo,
                certificado.data_execucao.strftime("%Y-%m-%d"),
                certificado.data_validade.strftime("%Y-%m-%d"),
                certificado.pragas_tratadas,
                certificado.arquivo_origem,
                certificado.data_cadastro.strftime("%Y-%m-%d %H:%M:%S"),
                certificado.valor or "",
                certificado.bairro or "",
                certificado.cidade or "",
            ]
        )
        
        for produto in bundle.produtos:
            valor = "" if produto.concentracao is None else f"{produto.concentracao:g}"
            produtos_sheet.append(
                [certificado.numero_certificado, produto.nome_produto, produto.classe_quimica, valor]
            )
        
        for metodo in bundle.metodos:
            metodos_sheet.append([certificado.numero_certificado, metodo.metodo, metodo.quantidade])
        
        workbook.save(self.consolidated_path)
        return self.consolidated_path
