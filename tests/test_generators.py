from __future__ import annotations

from pathlib import Path

import pytest

from engine_excel_to_pdf.generators.pdf_generator import PDFGenerator
from engine_excel_to_pdf.generators.spreadsheet_generator import SpreadsheetGenerator


class TestSpreadsheetGenerator:
    def test_generate_spreadsheet(self, temp_dir, sample_bundle, assets_dir):
        generator = SpreadsheetGenerator(output_dir=temp_dir)
        
        output_path = generator.generate(sample_bundle)
        
        assert output_path.exists()
        assert output_path.suffix == ".xlsx"
        assert output_path.name == "certificados_consolidados.xlsx"
    
    def test_spreadsheet_contains_data(self, temp_dir, sample_bundle):
        from openpyxl import load_workbook
        
        generator = SpreadsheetGenerator(output_dir=temp_dir)
        output_path = generator.generate(sample_bundle)
        
        wb = load_workbook(output_path)
        
        assert "certificado" in wb.sheetnames
        assert "produtos" in wb.sheetnames
        assert "metodos" in wb.sheetnames
        
        cert_sheet = wb["certificado"]
        assert cert_sheet.max_row == 2
    
    def test_spreadsheet_append_mode(self, temp_dir, sample_bundle):
        """Test that multiple calls append to the same file."""
        from openpyxl import load_workbook
        import copy
        
        generator = SpreadsheetGenerator(output_dir=temp_dir)
        
        output_path1 = generator.generate(sample_bundle)
        
        bundle2 = copy.deepcopy(sample_bundle)
        bundle2.certificado.numero_certificado = "CERT-2024-002"
        bundle2.certificado.id = None
        output_path2 = generator.generate(bundle2)
        
        assert output_path1 == output_path2
        
        wb = load_workbook(output_path2)
        cert_sheet = wb["certificado"]
        
        assert cert_sheet.max_row == 3


class TestPDFGenerator:
    def test_generate_pdf(self, temp_dir, sample_bundle, assets_dir):
        generator = PDFGenerator(
            output_dir=temp_dir,
            logo_path=None,
            template_name="certificado.html",
            stylesheet_name="certificado.css",
        )
        
        output_path = generator.generate(sample_bundle)
        
        assert output_path.exists()
        assert output_path.suffix == ".pdf"
        assert "CERT-2024-001" in output_path.name
    
    def test_pdf_file_size(self, temp_dir, sample_bundle, assets_dir):
        generator = PDFGenerator(output_dir=temp_dir)
        output_path = generator.generate(sample_bundle)
        
        assert output_path.stat().st_size > 100
