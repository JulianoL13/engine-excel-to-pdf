from __future__ import annotations

import pytest
from pathlib import Path
from openpyxl import Workbook

from engine_excel_to_pdf.extractor.excel_extractor import ExcelExtractor


class TestExcelExtractorFallback:
    """Test fallback mechanism for fields with variable cell locations."""
    
    def test_nome_fantasia_in_c17_primary_location(self, tmp_path):
        """Test extraction when nome_fantasia is in primary location C17."""
        # Create Excel with nome_fantasia in C17
        wb = Workbook()
        ws = wb.active
        
        # Fill required fields
        ws["C10"] = "CERT-2025-001"
        ws["I10"] = "LIC-MA-12345"
        ws["C15"] = "EMPRESA TESTE LTDA"
        ws["C17"] = "NOME FANTASIA EM C17"  # Primary location
        ws["D19"] = "Rua Teste, 123, Centro, São Paulo-MA"
        ws["E20"] = "11.222.333/0001-81"
        ws["E21"] = "15 DE JANEIRO DE 2025"
        ws["F22"] = "Insetos e roedores"
        ws["B48"] = "15 DE ABRIL DE 2025"
        
        # Save temporary file
        excel_file = tmp_path / "test_c17.xlsx"
        wb.save(excel_file)
        
        # Extract data
        extractor = ExcelExtractor()
        bundle = extractor.extract(excel_file)
        
        # Verify nome_fantasia was extracted from C17
        assert bundle.certificado.nome_fantasia == "NOME FANTASIA EM C17"
        assert bundle.certificado.razao_social == "EMPRESA TESTE LTDA"
    
    def test_nome_fantasia_in_d17_fallback_location(self, tmp_path):
        """Test extraction when nome_fantasia is in fallback location D17."""
        # Create Excel with nome_fantasia in D17 (C17 is empty)
        wb = Workbook()
        ws = wb.active
        
        # Fill required fields
        ws["C10"] = "CERT-2025-002"
        ws["I10"] = "LIC-MA-67890"
        ws["C15"] = "OUTRA EMPRESA LTDA"
        ws["C17"] = None  # Primary location is empty
        ws["D17"] = "NOME FANTASIA EM D17"  # Fallback location
        ws["D19"] = "Av. Principal, 456, Jardim, Imperatriz-MA"
        ws["E20"] = "22.333.444/0001-92"
        ws["E21"] = "20 DE FEVEREIRO DE 2025"
        ws["F22"] = "Pragas diversas"
        ws["B48"] = "20 DE MAIO DE 2025"
        
        # Save temporary file
        excel_file = tmp_path / "test_d17.xlsx"
        wb.save(excel_file)
        
        # Extract data
        extractor = ExcelExtractor()
        bundle = extractor.extract(excel_file)
        
        # Verify nome_fantasia was extracted from D17 (fallback)
        assert bundle.certificado.nome_fantasia == "NOME FANTASIA EM D17"
        assert bundle.certificado.razao_social == "OUTRA EMPRESA LTDA"
    
    def test_nome_fantasia_c17_priority_over_d17(self, tmp_path):
        """Test that C17 has priority when both C17 and D17 have values."""
        # Create Excel with nome_fantasia in both C17 and D17
        wb = Workbook()
        ws = wb.active
        
        # Fill required fields
        ws["C10"] = "CERT-2025-003"
        ws["I10"] = "LIC-MA-11111"
        ws["C15"] = "EMPRESA AMBOS LTDA"
        ws["C17"] = "VALOR EM C17"  # Primary (should be used)
        ws["D17"] = "VALOR EM D17"  # Fallback (should be ignored)
        ws["D19"] = "Rua Ambos, 789, Vila Nova, Carolina-MA"
        ws["E20"] = "33.444.555/0001-03"
        ws["E21"] = "10 DE MARÇO DE 2025"
        ws["F22"] = "Roedores"
        ws["B48"] = "10 DE JUNHO DE 2025"
        
        # Save temporary file
        excel_file = tmp_path / "test_both.xlsx"
        wb.save(excel_file)
        
        # Extract data
        extractor = ExcelExtractor()
        bundle = extractor.extract(excel_file)
        
        # Verify C17 has priority (primary location)
        assert bundle.certificado.nome_fantasia == "VALOR EM C17"
        assert bundle.certificado.numero_certificado == "CERT-2025-003"
    
    def test_nome_fantasia_empty_in_both_cells(self, tmp_path):
        """Test extraction when nome_fantasia is empty in both C17 and D17."""
        # Create Excel with nome_fantasia empty in both locations
        wb = Workbook()
        ws = wb.active
        
        # Fill required fields (except nome_fantasia)
        ws["C10"] = "CERT-2025-004"
        ws["I10"] = "LIC-MA-22222"
        ws["C15"] = "EMPRESA SEM FANTASIA LTDA"
        ws["C17"] = None  # Primary is empty
        ws["D17"] = None  # Fallback is also empty
        ws["D19"] = "Rua Vazia, 999, Centro, Imperatriz-MA"
        ws["E20"] = "44.555.666/0001-14"
        ws["E21"] = "25 DE ABRIL DE 2025"
        ws["F22"] = "Insetos"
        ws["B48"] = "25 DE JULHO DE 2025"
        
        # Save temporary file
        excel_file = tmp_path / "test_empty.xlsx"
        wb.save(excel_file)
        
        # Extract data
        extractor = ExcelExtractor()
        bundle = extractor.extract(excel_file)
        
        # Verify nome_fantasia is empty string (normalized)
        assert bundle.certificado.nome_fantasia == ""
        assert bundle.certificado.razao_social == "EMPRESA SEM FANTASIA LTDA"
    
    def test_other_fields_not_affected_by_fallback(self, tmp_path):
        """Test that fields without fallbacks still work normally."""
        # Create Excel with nome_fantasia in D17
        wb = Workbook()
        ws = wb.active
        
        # Fill all required fields
        ws["C10"] = "CERT-2025-005"
        ws["I10"] = "LIC-MA-33333"
        ws["C15"] = "EMPRESA NORMAL LTDA"
        ws["C17"] = None
        ws["D17"] = "FANTASIA FALLBACK"
        ws["D19"] = "Rua Normal, 111, Bairro, Cidade-MA"
        ws["E20"] = "55.666.777/0001-25"
        ws["E21"] = "30 DE MAIO DE 2025"
        ws["F22"] = "Pragas gerais"
        ws["B48"] = "30 DE AGOSTO DE 2025"
        
        # Save temporary file
        excel_file = tmp_path / "test_normal.xlsx"
        wb.save(excel_file)
        
        # Extract data
        extractor = ExcelExtractor()
        bundle = extractor.extract(excel_file)
        
        # Verify all other fields extracted correctly
        assert bundle.certificado.numero_certificado == "CERT-2025-005"
        assert bundle.certificado.numero_licenca == "LIC-MA-33333"
        assert bundle.certificado.razao_social == "EMPRESA NORMAL LTDA"
        assert bundle.certificado.nome_fantasia == "FANTASIA FALLBACK"
        assert bundle.certificado.cnpj == "55.666.777/0001-25"
        assert "Rua Normal" in bundle.certificado.endereco_completo
