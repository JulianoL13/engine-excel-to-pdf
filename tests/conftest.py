from __future__ import annotations

import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Generator

import pytest
from openpyxl import Workbook

from engine_excel_to_pdf.config import EngineConfig
from engine_excel_to_pdf.models import Certificado, CertificadoBundle, MetodoAplicacao, ProdutoQuimico


@pytest.fixture
def temp_dir() -> Generator[Path, None, None]:
    temp_path = Path(tempfile.mkdtemp())
    yield temp_path
    shutil.rmtree(temp_path, ignore_errors=True)


@pytest.fixture
def engine_config(temp_dir: Path) -> EngineConfig:
    return EngineConfig(output_dir=temp_dir)


@pytest.fixture
def sample_certificado() -> Certificado:
    return Certificado(
        numero_certificado="CERT-2024-001",
        numero_licenca="LIC-SP-12345",
        razao_social="Empresa Teste LTDA",
        nome_fantasia="Teste Corp",
        cnpj="11.222.333/0001-81",
        endereco_completo="Rua Teste, 123 - São Paulo/SP",
        data_execucao=date(2024, 1, 15),
        data_validade=date(2024, 7, 15),
        pragas_tratadas="Baratas, Formigas, Ratos",
        arquivo_origem="upload-excel",
        data_cadastro=datetime(2024, 1, 15, 10, 30, 0),
    )


@pytest.fixture
def sample_produtos() -> list[ProdutoQuimico]:
    return [
        ProdutoQuimico(
            nome_produto="Inseticida Alpha",
            classe_quimica="Piretroide",
            concentracao=2.5,
        ),
        ProdutoQuimico(
            nome_produto="Raticida Beta",
            classe_quimica="Anticoagulante",
            concentracao=0.005,
        ),
    ]


@pytest.fixture
def sample_metodos() -> list[MetodoAplicacao]:
    return [
        MetodoAplicacao(
            metodo="Pulverização",
            quantidade="500ml",
        ),
        MetodoAplicacao(
            metodo="Gel",
            quantidade="50g",
        ),
    ]


@pytest.fixture
def sample_bundle(
    sample_certificado: Certificado,
    sample_produtos: list[ProdutoQuimico],
    sample_metodos: list[MetodoAplicacao],
) -> CertificadoBundle:
    return CertificadoBundle(
        certificado=sample_certificado,
        produtos=sample_produtos,
        metodos=sample_metodos,
    )


@pytest.fixture
def sample_excel_file(temp_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    
    ws["C10"] = "CERT-2024-001"
    ws["I10"] = "LIC-SP-12345"
    ws["C15"] = "Empresa Teste LTDA"
    ws["C17"] = "Teste Corp"
    ws["D19"] = "Rua Teste, 123 - São Paulo/SP"
    ws["E20"] = "11.222.333/0001-81"
    ws["E21"] = "15 DE JANEIRO DE 2024"
    ws["F22"] = "Baratas, Formigas, Ratos"
    ws["B48"] = "15 DE JULHO DE 2024"
    
    ws.cell(row=25, column=6).value = "Piretroide"
    ws.cell(row=26, column=6).value = "Anticoagulante"
    
    ws.cell(row=29, column=4).value = "Inseticida Alpha"
    ws.cell(row=29, column=9).value = 2.5
    ws.cell(row=30, column=4).value = "Raticida Beta"
    ws.cell(row=30, column=9).value = 0.005
    
    ws.cell(row=34, column=4).value = "Pulverização"
    ws.cell(row=34, column=8).value = "500ml"
    ws.cell(row=35, column=4).value = "Gel"
    ws.cell(row=35, column=8).value = "50g"
    
    file_path = temp_dir / "test_certificado.xlsx"
    wb.save(file_path)
    return file_path


@pytest.fixture
def assets_dir(temp_dir: Path) -> Path:
    assets = temp_dir / "assets"
    templates = assets / "templates"
    templates.mkdir(parents=True)
    
    html_content = """<!DOCTYPE html>
<html>
<head><title>Certificado</title></head>
<body>
<h1>{{ certificate.numero_certificado }}</h1>
<p>{{ certificate.razao_social }}</p>
</body>
</html>"""
    
    (templates / "certificado.html").write_text(html_content)
    (templates / "certificado.css").write_text("body { font-family: Arial; }")
    
    return assets
