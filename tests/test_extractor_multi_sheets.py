from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from engine_excel_to_pdf.interface import MotorCertificados


def create_multi_sheet_file(tmpdir: Path) -> Path:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Vazia"
    ws2 = wb.create_sheet("Dados")

    ws2["C10"] = "CERT-2024-002"
    ws2["I10"] = "LIC-SP-67890"
    ws2["C15"] = "Empresa Multi LTDA"
    ws2["C17"] = "Multi Corp"
    ws2["D19"] = "Rua Multi, 789 - São Paulo/SP"
    ws2["E20"] = "11.222.333/0001-81"
    ws2["E21"] = "20 DE FEVEREIRO DE 2024"
    ws2["F22"] = "Mosquitos"
    ws2["B48"] = "20 DE AGOSTO DE 2024"

    ws2.cell(row=25, column=6).value = "Piretroide"
    ws2.cell(row=29, column=4).value = "Inseticida Multi"
    ws2.cell(row=29, column=9).value = 1.0
    ws2.cell(row=34, column=4).value = "Pulverização"
    ws2.cell(row=34, column=8).value = "300ml"

    file_path = tmpdir / "multi_sheets.xlsx"
    wb.save(file_path)
    return file_path


def test_extractor_selects_valid_sheet(engine_config, assets_dir, temp_dir):
    engine_config.assets_dir = assets_dir
    motor = MotorCertificados(config=engine_config)
    excel_path = create_multi_sheet_file(temp_dir)

    resultado = motor.processar_upload(excel_path)

    certificado = resultado["certificado"]
    assert certificado.numero_certificado == "CERT-2024-002"
    assert resultado["planilha"].exists()
    assert resultado["pdf"].exists()